#!/usr/bin/env python3
"""
extraer_glosas.py

Lee todos los PDFs de "NOTIFICACIÓN DE GLOSAS" (Salud Total EPS) que haya en una
carpeta, extrae la tabla de servicios glosados y consolida todo en un único Excel.

USO:
    python3 extraer_glosas.py /home/woombat/Descargas/proyectos/PDF_to_Excel_Mi_Amor/PDFs/pdf2026 /home/woombat/Descargas/proyectos/PDF_to_Excel_Mi_Amor/Excel_generado/excel_generado_2026.xlsx

Ejemplo:
    python3 extraer_glosas.py ./pdfs_glosas ./consolidado_glosas.xlsx

Reglas aplicadas (confirmadas con el usuario):
- PERIODO: si el mes de "Fecha Rad." está entre 1 y 6 -> "PRIMER"; entre 7 y 12 -> "SEGUNDO"
- MES: nombre del mes (en español, mayúsculas) tomado de "Fecha Rad."
- ESTADO PAGO: se deja en blanco (se diligencia manualmente después)
- Nombre reporte de glosa: nombre del archivo PDF de origen

NOTA TÉCNICA IMPORTANTE:
Salud Total no siempre usa exactamente las mismas columnas ni el mismo ancho de
tabla en sus cartas de glosa (por ejemplo, algunas cartas incluyen la columna
"Numreg" y otras no). Cuando la tabla es muy ancha, la última columna se desborda
a una página aparte como una "tabla" de una sola columna.

Por eso, en vez de asumir un número fijo de columnas, este script identifica cada
columna POR SU NOMBRE DE ENCABEZADO (usando MAPEO_ENCABEZADOS) y pega automáticamente
los fragmentos de columnas desbordadas con la fila principal a la que pertenecen,
usando el ancho de cada bloque de tabla como huella para saber a qué grupo de
columnas corresponde cada fragmento.
"""

import sys
import glob
import os
import re
import unicodedata
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}

# Columnas de salida, en el orden exacto solicitado
COLUMNAS_SALIDA = [
    "N#", "PERIODO", "MES", "Fecha Rad.", "Num Rad.", "Pref Fact.", "Num Fact.",
    "Num Doc Afil.", "Nombre Completo Afiliado", "Num Autoriz.", "Valor Servicio",
    "Cant. Fact.", "Valor Unitario", "Valor Total Glosa", "Valor Glosa Detalle",
    "ESTADO PAGO", "Descripcion Motivo", "OBSERVACIONES", "Cod Mot Glosa Genr",
    "Motivo Glosa General", "Cod Mot Glosa Espec", "Motivo Glosa Especifico",
    "Nombre reporte de glosa",
]

# Mapeo de encabezados tal como aparecen en el PDF (ya normalizados: sin tildes,
# minúsculas, espacios colapsados) hacia el nombre de campo interno que usamos.
# Si Salud Total agrega/quita columnas, casi siempre basta con agregar una entrada
# aquí para que el script se adapte, sin tocar el resto de la lógica.
MAPEO_ENCABEZADOS = {
    "fecha rad.": "Fecha Rad.",
    "num rad.": "Num Rad.",
    "pref fact.": "Pref Fact.",
    "num fact.": "Num Fact.",
    "numreg": "Numreg",
    "num doc afil.": "Num Doc Afil.",
    "nombre completo afili.": "Nombre Completo Afiliado",
    "num autoriz.": "Num Autoriz.",
    "servicio": "Servicio",
    "valor servicio": "Valor Servicio",
    "cant. fact.": "Cant. Fact.",
    "valor unitario": "Valor Unitario",
    "valor total glosa": "Valor Total Glosa",
    "valor glosa detalle": "Valor Glosa Detalle",
    "descripcion motivo": "Descripcion Motivo",
    "observaciones": "OBSERVACIONES",
    "cod mot glosa genr": "Cod Mot Glosa Genr",
    "motivo glosa general": "Motivo Glosa General",
    "cod mot glosa espec": "Cod Mot Glosa Espec",
    "motivo glosa especifico": "Motivo Glosa Especifico",
}

# Campos que identifican una fila de registro real (si ambos vienen vacíos,
# es una fila de totales/pie de tabla y se descarta)
CAMPOS_IDENTIFICADORES = ("Fecha Rad.", "Num Rad.")


def limpiar_texto(valor):
    """Quita saltos de línea y espacios repetidos dejados por pdfplumber."""
    if valor is None:
        return ""
    texto = str(valor).replace("\n", " ").strip()
    texto = re.sub(r"\s{2,}", " ", texto)
    return texto


def normalizar(texto):
    """minúsculas, sin tildes, espacios colapsados — para comparar encabezados."""
    texto = limpiar_texto(texto).lower()
    texto = "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )
    return texto


def parsear_fecha(fecha_str):
    """Convierte 'dd/mm/aaaa' a un objeto date. Devuelve None si no se puede."""
    fecha_str = limpiar_texto(fecha_str)
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(fecha_str, fmt).date()
        except ValueError:
            continue
    return None


def periodo_y_mes(fecha_rad):
    """Aplica la regla: mes 1-6 -> PRIMER, mes 7-12 -> SEGUNDO."""
    fecha = parsear_fecha(fecha_rad)
    if fecha is None:
        return "", ""
    periodo = "PRIMER" if fecha.month <= 6 else "SEGUNDO"
    mes = MESES_ES.get(fecha.month, "")
    return periodo, mes


def fila_es_encabezado(fila):
    """Una fila es encabezado si su primera celda es un nombre de campo conocido."""
    if not fila:
        return False
    return normalizar(fila[0]) in MAPEO_ENCABEZADOS


def mapear_fila(fila, campos):
    """Empareja cada celda de la fila con el nombre de campo correspondiente."""
    resultado = {}
    for nombre, valor in zip(campos, fila):
        if not nombre:
            continue
        resultado[nombre] = limpiar_texto(valor)
    return resultado


def extraer_filas_de_pdf(ruta_pdf, advertencias):
    """Devuelve una lista de diccionarios, una fila por servicio glosado."""
    nombre_archivo = os.path.basename(ruta_pdf)
    se_encontro_alguna_tabla = False

    filas_abiertas = []          # lista de dicts, una por registro real (no totales)
    encabezado_por_ancho = {}    # ancho de tabla -> lista de nombres de campo

    with pdfplumber.open(ruta_pdf) as pdf:
        for num_pagina, pagina in enumerate(pdf.pages, start=1):
            tablas = pagina.extract_tables()
            for tabla in tablas:
                if not tabla or not tabla[0]:
                    continue
                se_encontro_alguna_tabla = True
                filas_tabla = [[limpiar_texto(c) for c in fila] for fila in tabla]
                ancho = len(filas_tabla[0])
                primera_fila = filas_tabla[0]

                if fila_es_encabezado(primera_fila):
                    campos = [MAPEO_ENCABEZADOS.get(normalizar(c)) for c in primera_fila]
                    encabezado_por_ancho[ancho] = campos
                    filas_datos = filas_tabla[1:]
                else:
                    campos = encabezado_por_ancho.get(ancho)
                    if campos is None and encabezado_por_ancho:
                        # No hay un encabezado exacto para este ancho: se usa el más
                        # parecido conocido hasta ahora (mejor esfuerzo)
                        ancho_cercano = min(
                            encabezado_por_ancho.keys(), key=lambda a: abs(a - ancho)
                        )
                        campos = encabezado_por_ancho[ancho_cercano]
                        advertencias.append(
                            f"[{nombre_archivo}] Página {num_pagina}: tabla de {ancho} "
                            f"columna(s) sin encabezado propio; se usó por aproximación "
                            f"el encabezado de {ancho_cercano} columna(s)."
                        )
                    if campos is None:
                        adelanto = " | ".join(primera_fila[:5])
                        advertencias.append(
                            f"[{nombre_archivo}] Página {num_pagina}: tabla IGNORADA, no "
                            f"se reconoce ningún encabezado para interpretarla. "
                            f"Inicio de la fila: {adelanto}"
                        )
                        continue
                    filas_datos = filas_tabla

                es_grupo_principal = any(c in campos for c in CAMPOS_IDENTIFICADORES)

                if es_grupo_principal:
                    for fila in filas_datos:
                        d = mapear_fila(fila, campos)
                        if not d.get("Fecha Rad.") and not d.get("Num Rad."):
                            continue  # fila de totales o vacía
                        filas_abiertas.append(d)
                else:
                    # Fragmento de columna(s) desbordadas a otra página: se pega a
                    # los registros más antiguos que aún no tengan estos campos
                    candidatos = [
                        f for f in filas_abiertas
                        if not any(c in f for c in campos if c)
                    ]
                    for i, fila in enumerate(filas_datos):
                        if not any(fila):
                            continue
                        if i >= len(candidatos):
                            advertencias.append(
                                f"[{nombre_archivo}] Página {num_pagina}: fragmento de "
                                f"columna(s) ({', '.join(c for c in campos if c)}) sin "
                                f"fila principal donde encajar (posible desalineación). "
                                f"Valor(es): {' | '.join(fila)}"
                            )
                            continue
                        d = mapear_fila(fila, campos)
                        candidatos[i].update(d)

    filas = []
    for d in filas_abiertas:
        fecha_rad = d.get("Fecha Rad.", "")
        num_rad = re.sub(r"\s+", "", d.get("Num Rad.", ""))
        # Corrige números que quedaron partidos por un salto de línea justo
        # después de un guión, ej. "34311- 2459719138"
        num_autoriz = re.sub(r"-\s+", "-", d.get("Num Autoriz.", ""))
        periodo, mes = periodo_y_mes(fecha_rad)

        campos_faltantes = [
            c for c in ("Descripcion Motivo", "Motivo Glosa General", "Motivo Glosa Especifico")
            if not d.get(c)
        ]
        if campos_faltantes:
            advertencias.append(
                f"[{nombre_archivo}] Fila con Num Rad. {num_rad or '(vacío)'} / Num Fact. "
                f"{d.get('Num Fact.', '(vacío)')}: no se pudieron completar estos campos "
                f"(posible fragmento perdido o desalineado): {', '.join(campos_faltantes)}"
            )

        filas.append({
            "PERIODO": periodo,
            "MES": mes,
            "Fecha Rad.": fecha_rad,
            "Num Rad.": num_rad,
            "Pref Fact.": d.get("Pref Fact.", ""),
            "Num Fact.": d.get("Num Fact.", ""),
            "Num Doc Afil.": d.get("Num Doc Afil.", ""),
            "Nombre Completo Afiliado": d.get("Nombre Completo Afiliado", ""),
            "Num Autoriz.": num_autoriz,
            "Valor Servicio": d.get("Valor Servicio", ""),
            "Cant. Fact.": d.get("Cant. Fact.", ""),
            "Valor Unitario": d.get("Valor Unitario", ""),
            "Valor Total Glosa": d.get("Valor Total Glosa", ""),
            "Valor Glosa Detalle": d.get("Valor Glosa Detalle", ""),
            "ESTADO PAGO": "",
            "Descripcion Motivo": d.get("Descripcion Motivo", ""),
            "OBSERVACIONES": d.get("OBSERVACIONES", ""),
            "Cod Mot Glosa Genr": d.get("Cod Mot Glosa Genr", ""),
            "Motivo Glosa General": d.get("Motivo Glosa General", ""),
            "Cod Mot Glosa Espec": d.get("Cod Mot Glosa Espec", ""),
            "Motivo Glosa Especifico": d.get("Motivo Glosa Especifico", ""),
            "Nombre reporte de glosa": nombre_archivo,
        })

    if not se_encontro_alguna_tabla:
        advertencias.append(
            f"[{nombre_archivo}] PDF IGNORADO POR COMPLETO: no se detectó ninguna tabla "
            f"en el documento (puede ser un PDF escaneado/imagen, o con un formato distinto)."
        )
    elif not filas:
        advertencias.append(
            f"[{nombre_archivo}] Se encontró tabla pero 0 filas de datos válidas "
            f"(revisar si el formato de esa tabla es distinto al esperado)."
        )

    return filas


def a_numero(valor):
    """Intenta convertir un texto tipo '144000.00' o '$ 144000' a float."""
    if valor is None:
        return None
    texto = str(valor).replace("$", "").replace(",", "").strip()
    if texto == "":
        return None
    try:
        return float(texto)
    except ValueError:
        return None


def construir_excel(filas, archivo_salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Glosas"

    fuente_normal = Font(name="Arial", size=10)
    fuente_encabezado = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    relleno_encabezado = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for col_idx, nombre_col in enumerate(COLUMNAS_SALIDA, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre_col)
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    columnas_numericas = {
        "Valor Servicio", "Cant. Fact.", "Valor Unitario",
        "Valor Total Glosa", "Valor Glosa Detalle",
    }

    for i, fila in enumerate(filas, start=1):
        fila_out = {"N#": i, **fila}
        for col_idx, nombre_col in enumerate(COLUMNAS_SALIDA, start=1):
            valor = fila_out.get(nombre_col, "")
            if nombre_col in columnas_numericas:
                num = a_numero(valor)
                valor = num if num is not None else valor
            celda = ws.cell(row=i + 1, column=col_idx, value=valor)
            celda.font = fuente_normal
            celda.alignment = Alignment(vertical="top", wrap_text=(nombre_col in
                                         ("Descripcion Motivo", "Nombre Completo Afiliado",
                                          "Motivo Glosa Especifico", "OBSERVACIONES")))
            if nombre_col in columnas_numericas and isinstance(valor, float):
                celda.number_format = "#,##0.00"

    anchos = {
        "N#": 5, "PERIODO": 10, "MES": 11, "Fecha Rad.": 12, "Num Rad.": 17,
        "Pref Fact.": 9, "Num Fact.": 10, "Num Doc Afil.": 13,
        "Nombre Completo Afiliado": 26, "Num Autoriz.": 14, "Valor Servicio": 13,
        "Cant. Fact.": 9, "Valor Unitario": 13, "Valor Total Glosa": 14,
        "Valor Glosa Detalle": 14, "ESTADO PAGO": 12, "Descripcion Motivo": 45,
        "OBSERVACIONES": 18, "Cod Mot Glosa Genr": 10, "Motivo Glosa General": 15,
        "Cod Mot Glosa Espec": 10, "Motivo Glosa Especifico": 22,
        "Nombre reporte de glosa": 40,
    }
    for col_idx, nombre_col in enumerate(COLUMNAS_SALIDA, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(nombre_col, 15)

    wb.save(archivo_salida)


def main():
    if len(sys.argv) < 2:
        print("Uso: python3 extraer_glosas.py <carpeta_con_pdfs> [archivo_salida.xlsx]")
        sys.exit(1)

    carpeta = sys.argv[1]
    archivo_salida = sys.argv[2] if len(sys.argv) > 2 else "consolidado_glosas.xlsx"
    if not archivo_salida.lower().endswith(".xlsx"):
        archivo_salida += ".xlsx"

    pdfs = sorted(glob.glob(os.path.join(carpeta, "*.pdf")))
    if not pdfs:
        print(f"No se encontraron PDFs en: {carpeta}")
        sys.exit(1)

    todas_las_filas = []
    advertencias = []
    for ruta_pdf in pdfs:
        print(f"Procesando: {os.path.basename(ruta_pdf)}")
        try:
            filas = extraer_filas_de_pdf(ruta_pdf, advertencias)
        except Exception as e:
            advertencias.append(
                f"[{os.path.basename(ruta_pdf)}] ERROR INESPERADO, el PDF se saltó por "
                f"completo: {type(e).__name__}: {e}"
            )
            print(f"  -> ERROR: {e} (se registró en advertencias y se continúa con los demás)")
            continue
        print(f"  -> {len(filas)} fila(s) extraída(s)")
        todas_las_filas.extend(filas)

    construir_excel(todas_las_filas, archivo_salida)
    print(f"\nListo. {len(todas_las_filas)} fila(s) en total guardadas en: {archivo_salida}")

    if advertencias:
        print(f"\n⚠ ADVERTENCIAS ({len(advertencias)}) — revisa lo siguiente manualmente:")
        for adv in advertencias:
            print(f"  - {adv}")

        archivo_log = os.path.splitext(archivo_salida)[0] + "_advertencias.txt"
        with open(archivo_log, "w", encoding="utf-8") as f:
            f.write(f"Advertencias generadas al procesar la carpeta: {carpeta}\n")
            f.write(f"Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            for adv in advertencias:
                f.write(f"- {adv}\n")
        print(f"\n(También quedaron guardadas en: {archivo_log})")
    else:
        print("\nSin advertencias: todos los PDFs y filas se procesaron correctamente.")


if __name__ == "__main__":
    main()